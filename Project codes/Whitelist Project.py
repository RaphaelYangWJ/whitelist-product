# === Import Modules
import pandas as pd
import matplotlib.pyplot as plt
import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter.ttk import *
from tkinter.filedialog import askdirectory
from tkinter.filedialog import askopenfilename
import ttkbootstrap as ttk
import tkinter.font as tf
import math
import pandas as pd
import numpy as np
from scipy.stats import norm
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans
from mpl_toolkits.axes_grid1.inset_locator import inset_axes
from matplotlib import ticker
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import seaborn as sns
import os
from openpyxl import load_workbook
from datetime import datetime

# === Class: threshold Algorithm
class Threshold_Algorithm(object):
    
    # ******************* Core Functions *******************
    # === Back-End Function: 定义初始参数
    def __init__(self):
        self.k_no_value = 3
        self.ci_value = 80
        self.yr_range_value = 0.3
        self.param_set = 1
        self.export_set = 1 
        self.output_fig = None
        self.theme = {
                        "White":["white","black","steelblue","gold","red","darkgoldenrod","red","w","k","indigo","Blues"],
                        "Black":["black","white","c","yellow","red","yellow","gold","k","w","peachpuff","crest"]
                        
                    }
        self.theme_select = None
    # === Function: 数据导入 --> 读取返回所有表格：货物贸易，货物贸易月度，货物贸易年度，服务贸易，服务贸易月度，服务贸易年度
    def data_import(self,goods_path,service_path):
        # === 货物贸易
        goods = pd.read_excel(goods_path,skiprows = 3)
        goods.set_index("货物贸易监测指标",inplace=True,drop=True)
        # goods.drop("Status",inplace=True,axis=1)
        # goods = goods.transpose()
        goods = goods.replace("N/A",None)
        for i in goods.columns.tolist():
            goods[i] = goods[i].apply(lambda x: str(x).replace(",","") if "," in str(x) else x)
            goods[i] = goods[i].apply(lambda x: round(float(str(x).replace("%",""))/100,2) if "%" in str(x) else x)
            goods[i] = goods[i].apply(lambda x: float(x))
        # === 获取数据指标
        goods_metrics = [] 
        value_metrics = []
        for i in goods.columns.tolist():
            if "MOM" in i or "YOY" in i:
                goods_metrics.append(i)
            else:
                value_metrics.append(i)
        # 生成指标和热力图数据集
        goods_ratio = goods[goods_metrics]
        self.heap_goods = goods.copy()[value_metrics]
        # === 判断是月度还是滚动年度指标
        self.good_frequency = "Year" if "-" in str([i for i in goods.index][1]) else "Month"
            
        
        
        # === 服务贸易
        service = pd.read_excel(service_path,skiprows = 3)
        service.set_index("服务贸易监测指标",inplace=True,drop=True)
        # service.drop("Status",inplace=True,axis=1)
        # service = service.transpose()
        service = service.replace("N/A",None)
        for i in service.columns.tolist():
            service[i] = service[i].apply(lambda x: str(x).replace(",","") if "," in str(x) else x)
            service[i] = service[i].apply(lambda x: round(float(str(x).replace("%",""))/100,2) if "%" in str(x) else x)
            service[i] = service[i].apply(lambda x: float(x))
        # === 获取数据指标
        service_metrics = [] 
        value_metrics = []
        for i in service.columns.tolist():
            if "MOM" in i or "YOY" in i:
                service_metrics.append(i)
            else:
                value_metrics.append(i)
        # 生成指标和热力图数据集
        service_ratio = service[service_metrics]
        self.heap_service = service.copy()[value_metrics]
        # === 判断是月度还是滚动年度指标
        self.service_frequency = "Year" if "-" in str([i for i in service.index][1]) else "Month"

        return goods,goods_ratio,service,service_ratio

    # === Function: 指标设置 - K-Means 机器学习 与 正态分布映射算法
    def K_Means_Algorithm(self, indicator=None, ax_left=0, n_cate=0, yellow_red_range=0, prob=0, source_data=None, line_color=None):
        # ***** K-Means 计算方法 *****
        # === 创建空的Threshold Panel
        Threshold_result = pd.DataFrame(columns = ["Indicator","Upper_red","Upper_yellow","Lower_yellow","Lower_red","Sus_hit","Risk_hit"])
        # === 对数据点去除极值 - 4个标准差
        overall_std = source_data[indicator].std() * 4
        overall_mean = source_data[indicator].mean()
        indicator_value = [i for i in source_data[indicator]]
        for i in range(0,len(indicator_value)):
            if (indicator_value[i] > (overall_mean + overall_std)) or (indicator_value[i] < (overall_mean - overall_std)):
                indicator_value[i] = None
            else: None
        source_data[indicator] = indicator_value
        # === 判断数据点是否充分
        if len(source_data[indicator].drop_duplicates()) <= 4:
            # === 计算K-Means
            source_data[indicator] = source_data[indicator].fillna(0)
            data_a = [i for i in source_data[indicator]]
            data = list(zip(data_a))
            kmeans = KMeans(n_clusters=1, n_init=10)
            kmeans.fit(data)
        else:
            # === 计算K-Means
            source_data[indicator] = source_data[indicator].fillna(0)
            data_a = [i for i in source_data[indicator]]
            data = list(zip(data_a))
            kmeans = KMeans(n_clusters=n_cate, n_init=10)
            kmeans.fit(data)
        # === 将K-Means聚类结果加入dataframe
        temp_record = pd.DataFrame(columns=[indicator,indicator+"_K Means Label"])
        temp_record[indicator] = data_a
        temp_record[indicator+"_K Means Label"] = kmeans.labels_.tolist()
        temp_process = temp_record.copy()
        # === 获取最大值和最小值
        min_value = temp_process[indicator].min()
        max_value = temp_process[indicator].max()
        # === 获取上下界标签
        min_group_label = temp_process.loc[temp_process[indicator] == min_value,indicator+"_K Means Label"].values[0]
        max_group_label = temp_process.loc[temp_process[indicator] == max_value,indicator+"_K Means Label"].values[0]
        # === 计算上下边界聚类
        lower_class = temp_process[temp_process[indicator+"_K Means Label"] == min_group_label]
        lower_class = np.array([float(i) for i in lower_class[indicator]])
        upper_class = temp_process[temp_process[indicator+"_K Means Label"] == max_group_label]
        upper_class = np.array([float(i) for i in upper_class[indicator]])
        
        # ***** 阈值算法 - 代码开始处*****
        # === 非加权算法
        overall_mean = np.mean(temp_process[indicator])
        overall_std = np.std(temp_process[indicator])

        # === 计算标准正态分布的分位数（给定概率计算)
        nd_quant_value, _= norm.interval(prob/100)
        nd_quant_value = abs(nd_quant_value)
        # === 计算上线边界
        if max_value / overall_mean > 40:
            upper_yellow_thres = (overall_mean) + nd_quant_value * (np.std(upper_class) + overall_std)
            upper_red_thres = upper_yellow_thres + (yellow_red_range + nd_quant_value) * (np.std(upper_class) + overall_std)
        else:
            upper_yellow_thres = (np.mean(upper_class) + overall_mean*2)/3 + nd_quant_value * (np.std(upper_class) + overall_std)
            upper_red_thres = upper_yellow_thres + (yellow_red_range + nd_quant_value) * (np.std(upper_class) + overall_std)
        # === 计算下线边界
        if max_value / overall_mean > 40:
            lower_yellow_thres = (np.mean(lower_class) + overall_mean*2)/3 - nd_quant_value * (np.std(lower_class) + overall_std)
            lower_red_thres = lower_yellow_thres - (yellow_red_range + nd_quant_value) * (np.std(lower_class) + overall_std)
        else:
            lower_yellow_thres = (np.mean(lower_class) + overall_mean*2)/3 - nd_quant_value * (np.std(lower_class) + overall_std)
            lower_red_thres = lower_yellow_thres - (yellow_red_range + nd_quant_value) * (np.std(lower_class) + overall_std)

        # === 对下线黄线做处理 - 不能大于0
        if lower_yellow_thres > 0:
            lower_yellow_thres = -0.1
        # === 将阈值扩展到百分比
        lower_yellow_thres = round(lower_yellow_thres,2)
        upper_yellow_thres = round(upper_yellow_thres,2)
        lower_red_thres = round(lower_red_thres,2)
        upper_red_thres = round(upper_red_thres,2)

        # === 所有数据都按照0个位数取整
        def Five_Zero(threshold):
            process_ts = int(round(threshold * 100))
            pos_neg_sign = 1 if threshold >= 0 else -1
            final_digit = int(str(process_ts)[-1])
            if final_digit == 1:
                return (process_ts - (1 * pos_neg_sign))/100
            elif final_digit == 2:
                return (process_ts - (2 * pos_neg_sign))/100
            elif final_digit == 3:
                return (process_ts - (3 * pos_neg_sign))/100
            elif final_digit == 4:
                return (process_ts - (4 * pos_neg_sign))/100
            elif final_digit == 5:
                return (process_ts + (5 * pos_neg_sign))/100
            elif final_digit == 6:
                return (process_ts + (4 * pos_neg_sign))/100
            elif final_digit == 7:
                return (process_ts + (3 * pos_neg_sign))/100
            elif final_digit == 8:
                return (process_ts + (2 * pos_neg_sign))/100
            elif final_digit == 9:
                return (process_ts + (1 * pos_neg_sign))/100
            else:
                return process_ts/100
        lower_yellow_thres = Five_Zero(lower_yellow_thres)
        upper_yellow_thres = Five_Zero(upper_yellow_thres)
        lower_red_thres = Five_Zero(lower_red_thres)
        upper_red_thres = Five_Zero(upper_red_thres)

        # === 对于小于0.05的值统一放在0.05
        if lower_yellow_thres >= -0.05:
            lower_yellow_thres = -0.05
            lower_red_thres = -0.1
        elif lower_yellow_thres >= -0.1:
            lower_yellow_thres = -0.1
            lower_red_thres = -0.15
        if upper_yellow_thres <= 0.05:
            upper_yellow_thres = 0.05
            upper_red_thres = 0.1
        elif upper_yellow_thres <= 0.1:
            upper_yellow_thres = 0.1
            upper_red_thres = 0.15
        
        # === 将相近结果区分开
        if abs(upper_red_thres - upper_yellow_thres) < 0.01:
            upper_red_thres = upper_red_thres + 0.05
        if abs(lower_red_thres - lower_yellow_thres) < 1:
            lower_red_thres = lower_red_thres - 0.05
        
        # === 对下线红线与做处理 - 不能小于100%和大于黄线
        if lower_red_thres <= -1:
            lower_red_thres = -1.01
        if lower_yellow_thres <= lower_red_thres:
            lower_yellow_thres = -0.9
            
        # === 缺乏历史数据时设置初始化指标 (所有历史数据均为0)
        if np.std(temp_process[indicator]) == 0:
            lower_yellow_thres = -0.05
            lower_red_thres = -0.10
            upper_yellow_thres = 0.05
            upper_red_thres = 0.10
        
        # ***** 阈值算法 - 代码结束处*****        
        # === 获取边界击中次数并年化
        risk_hit = len([i for i in temp_record[indicator] if i <= lower_red_thres]) + len([i for i in temp_record[indicator] if i >= upper_red_thres])
        sus_hit = len([i for i in temp_record[indicator] if i <= lower_yellow_thres]) + len([i for i in temp_record[indicator] if i >= upper_yellow_thres]) - risk_hit
        if risk_hit != 0:
            annualize_risk_hit = round((risk_hit/len([i for i in temp_record[indicator]])) * 12,2)
        else:
            annualize_risk_hit = 0
        if sus_hit != 0:
            annualize_sus_hit = round((sus_hit/len([i for i in temp_record[indicator]])) * 12,2)
        else:
            annualize_sus_hit = 0

        # === 将计算结果输出到Panel
        Threshold_result.loc[len(Threshold_result)+1] = [indicator,upper_red_thres,upper_yellow_thres,lower_yellow_thres,lower_red_thres,annualize_sus_hit,annualize_risk_hit]
        
        # ***** 数据可视化 *****
        # === 可视化数据计算结果
        temp_df = temp_process.copy()
        temp_df["time"] = [i for i in range(0,len(temp_df))]
        Lower_yellow = Threshold_result.loc[Threshold_result["Indicator"]==indicator,"Lower_yellow"].values[0]
        Lower_red = Threshold_result.loc[Threshold_result["Indicator"]==indicator,"Lower_red"].values[0]
        Upper_yellow = Threshold_result.loc[Threshold_result["Indicator"]==indicator,"Upper_yellow"].values[0]
        Upper_red = Threshold_result.loc[Threshold_result["Indicator"]==indicator,"Upper_red"].values[0]
        temp_df["Lower_yellow"] = Lower_yellow
        temp_df["Lower_red"] = Lower_red
        temp_df["Upper_yellow"] = Upper_yellow
        temp_df["Upper_red"] = Upper_red
        # === 建立三幅图及比例确定
        ax_main = inset_axes(ax_left,width="70%",height="80%",loc=2)
        ax_right = inset_axes(ax_left,width="25%",height="95%",loc=1)
        ax_bottom = inset_axes(ax_left,width="70%",height="15%",loc=3)
        # === 设置背景色
        ax_right.set_facecolor(self.theme[self.theme_select][0])
        ax_bottom.set_facecolor(self.theme[self.theme_select][0])
        ax_main.set_facecolor(self.theme[self.theme_select][0])
        
        # === 对主图绘图
        ax_main.plot(temp_df["time"],temp_df[indicator],color=line_color, marker=".",label=indicator,linewidth=2)
        ax_main.plot(temp_df["time"],temp_df["Lower_yellow"],color=self.theme[self.theme_select][3],label="黄色下线预警线",linewidth=2)
        ax_main.plot(temp_df["time"],temp_df["Upper_yellow"],color=self.theme[self.theme_select][3],label="黄色上线预警线",linewidth=2)
        ax_main.plot(temp_df["time"],temp_df["Lower_red"],color=self.theme[self.theme_select][4],label="红色下线预警线",linewidth=2)
        ax_main.plot(temp_df["time"],temp_df["Upper_red"],color=self.theme[self.theme_select][4],label="红色上线预警线",linewidth=2)
        ax_main.grid(axis="y",linestyle="--")
        # === 显示threshold 数据点 - 在最后一位显示
        ax_main.text(temp_df["time"].to_list()[0],temp_df["Lower_yellow"].to_list()[0],str(round(temp_df["Lower_yellow"].to_list()[0]*100))+"%",ha="right",va="top",fontsize=10, color=self.theme[self.theme_select][5],bbox=dict(boxstyle="round",fc=self.theme[self.theme_select][7],edgecolor=self.theme[self.theme_select][8],alpha=0.9))
        ax_main.text(temp_df["time"].to_list()[0],temp_df["Upper_yellow"].to_list()[0],str(round(temp_df["Upper_yellow"].to_list()[0]*100))+"%",ha="right",va="bottom",fontsize=10, color=self.theme[self.theme_select][5],bbox=dict(boxstyle="round",fc=self.theme[self.theme_select][7],edgecolor=self.theme[self.theme_select][8],alpha=0.9))
        ax_main.text(temp_df["time"].to_list()[-1],temp_df["Lower_red"].to_list()[-1],str(round(temp_df["Lower_red"].to_list()[-1]*100))+"%",ha="right",va="bottom",fontsize=10,color=self.theme[self.theme_select][6],bbox=dict(boxstyle="round",fc=self.theme[self.theme_select][7],edgecolor=self.theme[self.theme_select][8],alpha=0.9))
        ax_main.text(temp_df["time"].to_list()[-1],temp_df["Upper_red"].to_list()[-1],str(round(temp_df["Upper_red"].to_list()[-1]*100))+"%",ha="right",va="top",fontsize=10,color=self.theme[self.theme_select][6],bbox=dict(boxstyle="round",fc=self.theme[self.theme_select][7],edgecolor=self.theme[self.theme_select][8],alpha=0.9))

        # === 显示数据点
        date_label = [i for i in temp_df["time"]]
        value_label = [i for i in temp_df[indicator]]
        value_true_label= [str(round(i*100,1))+"%" for i in temp_df[indicator]]
        for a,b,c in zip(date_label,value_label,value_true_label):
            ax_main.text(a,b,c,ha="right",va="bottom",fontsize=9,color=line_color)
        ax_main.yaxis.set_major_formatter(ticker.PercentFormatter(xmax=1,decimals=0))
        ax_main.set_fontsize = (12)
        ax_main.grid(linestyle = '--')
        ax_main.set_xticks([])
        # === 对右侧图绘图
        ax_right.scatter(temp_df["time"],temp_df[indicator],c=kmeans.labels_,cmap="Set2",s=50) # 透明度
        ax_right.set_fontsize = (12)
        ax_right.set_xticks([])
        ax_right.set_yticks([])
        # === 对底部图绘图
        ax_bottom.hist(temp_df[indicator],  color='cadetblue',rwidth=0.5)
        ax_bottom.set_fontsize = (12)
        x = np.linspace(Lower_red-0.2,Upper_red+0.2,1000)
        fx = (1 / (overall_std * (2 * np.pi)**0.5) * np.exp(-(x - overall_mean)**2 / (2 * overall_std**2)))*20
        ax_bottom.plot(x,fx,color=self.theme[self.theme_select][2])
        ax_bottom.axvline(Lower_yellow, color=self.theme[self.theme_select][5],linewidth=2)
        ax_bottom.axvline(Lower_red, color=self.theme[self.theme_select][6],linewidth=2)
        ax_bottom.axvline(Upper_yellow, color=self.theme[self.theme_select][5],linewidth=2)
        ax_bottom.axvline(Upper_red, color=self.theme[self.theme_select][6],linewidth=2)
        ax_bottom.set_xticks([])

        ax_main.spines["top"].set_color(self.theme[self.theme_select][1])
        ax_main.spines["left"].set_color(self.theme[self.theme_select][1])
        ax_main.spines["right"].set_color(self.theme[self.theme_select][1])
        ax_main.spines["bottom"].set_color(self.theme[self.theme_select][1])
        
        ax_right.spines["top"].set_color(self.theme[self.theme_select][1])
        ax_right.spines["left"].set_color(self.theme[self.theme_select][1])
        ax_right.spines["right"].set_color(self.theme[self.theme_select][1])
        ax_right.spines["bottom"].set_color(self.theme[self.theme_select][1])
        ax_bottom.spines["top"].set_color(self.theme[self.theme_select][1])
        ax_bottom.spines["left"].set_color(self.theme[self.theme_select][1])
        ax_bottom.spines["right"].set_color(self.theme[self.theme_select][1])
        ax_bottom.spines["bottom"].set_color(self.theme[self.theme_select][1])

        ax_main.tick_params(axis='x', colors=self.theme[self.theme_select][1])
        ax_main.tick_params(axis='y', colors=self.theme[self.theme_select][1])
        ax_right.tick_params(axis='x', colors=self.theme[self.theme_select][1])
        ax_right.tick_params(axis='y', colors=self.theme[self.theme_select][1])
        ax_bottom.tick_params(axis='x', colors=self.theme[self.theme_select][1])
        ax_bottom.tick_params(axis='y', colors=self.theme[self.theme_select][1])

        return Threshold_result, ax_main, temp_df, temp_df
    # === Function: 所有指标计算与绘图
    def All_Metrics_Computations(self,good_data,service_data,figure_canvas,n_cate,yellow_red_range,prob):
        # 创建变量
        self.hit_record = {}
        self.name_dict = {}
        name_count = 1
        # 清空所有子图
        figure_canvas.clear()
        # 创建变量 - 绘图定位数据记录
        self.visual_record = pd.DataFrame(columns=["indicator","row","col"])
        # 定义中文输出
        # self.amend_Ops_text("中文输出定义完成")
        plt.rcParams['font.sans-serif']=['Microsoft YaHei']
        plt.rcParams['axes.unicode_minus'] = False
        plt.clf()
        self.amend_Ops_text("Completed for cleaning canvas")
        self.amend_Ops_text("Start to train K-Means model for param setup")
        # === 构建结果列表
        result_trade = pd.DataFrame(columns = ["Indicator","Upper_red","Upper_yellow","Lower_yellow","Lower_red","Sus_hit","Risk_hit"])
        # ********************* 绘图函数 ***************************
        # 计算货物贸易与服务贸易的总共数量
        goods_col = good_data.columns.tolist()
        service_col = service_data.columns.tolist()
        # 规定每行显示的图表数量
        col_show = 4
        total_len = len(goods_col) + len(service_col) # 获取总长度
        # 计算所需要的列
        row_show = math.ceil(total_len / col_show)+12
        # 初始化row coordinate 和 col coordinate 以及grid 构图
        row_coordinate, col_coordinate = 0,0
        grid_panel = plt.GridSpec(row_show, col_show, hspace=0.5, wspace=0.2)
        # 计算进度
        each_step = (80-40)/total_len
        current_pro = 40
        # === 绘制货物贸易图
        if len(goods_col) > 0:
            for i in goods_col:
                ax = figure_canvas.add_subplot(grid_panel[row_coordinate:row_coordinate+1, col_coordinate:col_coordinate+1])
                ax.set_title("Goods Trade： "+i + "\nK-Means Learning Result_Code: "+str(name_count),fontdict={"color":self.theme[self.theme_select][1]})
                ax.axis('off')
                temp,globals()["ax_main"+str(name_count)],globals()["temp_df"+str(name_count)],data_df = self.K_Means_Algorithm(indicator = i,ax_left=ax,n_cate=n_cate,yellow_red_range=yellow_red_range,prob=prob,source_data=good_data,line_color=self.theme[self.theme_select][2])
                self.name_dict[str(name_count)+"-货物贸易-" + i] = str(name_count)
                temp["Indicator"] = temp["Indicator"].apply(lambda x: str(name_count)+"-货物贸易-" + x)
                result_trade = pd.concat([result_trade,temp])
                # 最近一年的击中情况放入字典中
                latest_temp = data_df.copy().tail(12)
                latest_temp["hit"] = 0
                latest_temp["hit"] = latest_temp[i].apply(lambda x: 1 if x > [i for i in latest_temp["Upper_yellow"]][0] else 0)
                latest_temp["hit"] = latest_temp[i].apply(lambda x: 1 if x < [i for i in latest_temp["Lower_yellow"]][0] else 0)
                hit_list = [i for i in latest_temp["hit"]]
                if len(hit_list) < 12:
                    hit_list = (12-len(hit_list)) * [0] + hit_list
                self.hit_record[str(name_count)+"-货物贸易-"+i] = hit_list
                # 将绘图记录放入visual record中
                self.visual_record.loc[len(self.visual_record)+1] = [str(name_count)+"-货物贸易-" + i,row_coordinate,col_coordinate]
                name_count += 1
                # 行列重新计算 - 智能布局
                if col_coordinate < (col_show-1):
                    col_coordinate += 1
                elif col_coordinate == (col_show-1):
                    row_coordinate += 1
                    col_coordinate = 0
                self.amend_Ops_text("Completed：货物贸易 - "+i)
                current_pro += each_step
                self.progress_update(current_pro)
        row_coordinate = row_coordinate + 1
        col_coordinate = 0
        # === 绘制服务贸易图
        if len(service_col) > 0:
            for i in service_col:
                ax = figure_canvas.add_subplot(grid_panel[row_coordinate:row_coordinate+1, col_coordinate:col_coordinate+1])
                ax.set_title("Service Trade： "+i + "\nK-Means Learning Result_Code: "+str(name_count),fontdict={"color":self.theme[self.theme_select][1]})
                ax.axis('off')
                temp,globals()["ax_main"+str(name_count)],globals()["temp_df"+str(name_count)],data_df= self.K_Means_Algorithm(indicator = i,ax_left=ax,n_cate=n_cate,yellow_red_range=yellow_red_range,prob=prob,source_data=service_data,line_color=self.theme[self.theme_select][9])
                self.name_dict[str(name_count)+"-服务贸易-" + i] = str(name_count)
                temp["Indicator"] = temp["Indicator"].apply(lambda x: str(name_count)+"-服务贸易-" + x)
                result_trade = pd.concat([result_trade,temp])
                # 最近一年的击中情况放入字典中
                latest_temp = data_df.copy().tail(12)
                latest_temp["hit"] = 0
                latest_temp["hit"] = latest_temp[i].apply(lambda x: 1 if x > [i for i in latest_temp["Upper_yellow"]][0] else 0)
                latest_temp["hit"] = latest_temp[i].apply(lambda x: 1 if x < [i for i in latest_temp["Lower_yellow"]][0] else 0)
                hit_list = [i for i in latest_temp["hit"]]
                if len(hit_list) < 12:
                    hit_list = (12-len(hit_list)) * [0] + hit_list
                self.hit_record[str(name_count)+"-货物贸易-"+i] = hit_list
                # 将绘图记录放入visual record中
                self.visual_record.loc[len(self.visual_record)+1] = [str(name_count)+"-服务贸易-" + i,row_coordinate,col_coordinate]
                name_count += 1
                # 行列重新计算 - 智能布局
                if col_coordinate < (col_show-1):
                    col_coordinate += 1
                elif col_coordinate == (col_show-1):
                    row_coordinate += 1
                    col_coordinate = 0
                self.amend_Ops_text("Completed：货物贸易 - "+i)
                current_pro += each_step
                self.progress_update(current_pro)
        self.amend_Ops_text("Compute the Heatmap")
        # === 加入相关性矩阵图
        ax_hm_1 = figure_canvas.add_subplot(grid_panel[row_coordinate+1:row_coordinate+5, 0:4])
        ax_hm_1.set_title("Correlation Matrix Heat Map - Goods",fontdict={"color":self.theme[self.theme_select][1]})
        ax_hm_1.set_facecolor(self.theme[self.theme_select][0])
        ax_hm_1.tick_params(axis='x', colors=self.theme[self.theme_select][1])
        ax_hm_1.tick_params(axis='y', colors=self.theme[self.theme_select][1])
        
        used_col = []
        for i in self.heap_goods.columns.to_list():
            if self.heap_goods[i].sum()!=0:
                used_col.append(i)
        input_matrix_copy = self.heap_goods[used_col]
        if len(input_matrix_copy.columns.to_list()) > 2:
            input_matrix_copy = input_matrix_copy.fillna(0)
            input_matrix_copy = input_matrix_copy.corr()
            sns.heatmap(input_matrix_copy, annot=True, vmax=1, square=True, cmap=self.theme[self.theme_select][10],ax=ax_hm_1)
        
        # === 加入相关性矩阵图
        ax_hm_2 = figure_canvas.add_subplot(grid_panel[row_coordinate+6:row_coordinate+10, 0:4])
        ax_hm_2.set_title("Correlation Matrix Heat Map - Services",fontdict={"color":self.theme[self.theme_select][1]})
        ax_hm_2.set_facecolor(self.theme[self.theme_select][0])
        ax_hm_2.tick_params(axis='x', colors=self.theme[self.theme_select][1])
        ax_hm_2.tick_params(axis='y', colors=self.theme[self.theme_select][1])
        used_col = []
        for i in self.heap_service.columns.to_list():
            if self.heap_service[i].sum()!=0:
                used_col.append(i)
        input_matrix_copy = self.heap_service[used_col]
        if len(input_matrix_copy.columns.to_list()) > 2:
            input_matrix_copy = input_matrix_copy.fillna(0)
            input_matrix_copy = input_matrix_copy.corr()
            sns.heatmap(input_matrix_copy, annot=True, vmax=1, square=True, cmap=self.theme[self.theme_select][10],ax=ax_hm_2, linecolor="black")

        # === Canvas 绘图
        self.Latest_Hit_Count(self.hit_record)
        figure_canvas.suptitle("Whitelist Threshold Setup Recommendation Data Panel\n\n\n"+self.chn_name+"\n\n\n Scroll Down to View Visualized Result\n\n"+\
                                     '1 Year Hit Month Count: '+str(12 - self.hit_outcome), fontsize=26, color = self.theme[self.theme_select][1])
        self.output_fig = figure_canvas
        height_para = row_show * 450
        ini_canvas = tk.Canvas(self.window_2,width=2000,height=1150,scrollregion=(0,0,2100,height_para))
        ini_canvas.pack(fill=BOTH)
        dashboard = tk.Frame(ini_canvas)
        vbar=Scrollbar(ini_canvas,orient=VERTICAL)
        vbar.place(x=2000,height=1150,width=40)
        vbar.configure(command=ini_canvas.yview)
        ini_canvas.config(yscrollcommand=vbar.set)   
        ini_canvas.create_window((2000/2,height_para/2),window=dashboard,height=height_para,width=2000)
        self.canvas = FigureCanvasTkAgg(figure_canvas, master=dashboard) 
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        # 返回结果
        return result_trade
    # === Functions: 指标设置 - 回溯绘图
    def K_Means_Algorithm_Resume(self, indicator=None, ax_left=0, n_cate=0, yellow_red_range=0, prob=0, source_data=None, line_color=None,Threshold_result=None):
        # === 对数据点去除极值 - 5个标准差
        overall_std = source_data[indicator].std() * 3
        overall_mean = source_data[indicator].mean()
        indicator_value = [i for i in source_data[indicator]]
        # === 对数据点去除极值 - 4个标准差
        overall_std = source_data[indicator].std() * 4
        overall_mean = source_data[indicator].mean()
        indicator_value = [i for i in source_data[indicator]]
        for i in range(0,len(indicator_value)):
            if (indicator_value[i] > (overall_mean + overall_std)) or (indicator_value[i] < (overall_mean - overall_std)):
                indicator_value[i] = None
            else: None
        source_data[indicator] = indicator_value
        # === 判断数据点是否充分
        if len(source_data[indicator].drop_duplicates()) <= 4:
            # === 计算K-Means
            source_data[indicator] = source_data[indicator].fillna(0)
            data_a = [i for i in source_data[indicator]]
            data = list(zip(data_a))
            kmeans = KMeans(n_clusters=1, n_init=10)
            kmeans.fit(data)
        else:
            # === 计算K-Means
            source_data[indicator] = source_data[indicator].fillna(0)
            data_a = [i for i in source_data[indicator]]
            data = list(zip(data_a))
            kmeans = KMeans(n_clusters=n_cate, n_init=10)
            kmeans.fit(data)
        # === 将K-Means聚类结果加入dataframe
        temp_record = pd.DataFrame(columns=[indicator,indicator+"_K Means Label"])
        temp_record[indicator] = data_a
        temp_record[indicator+"_K Means Label"] = kmeans.labels_.tolist()
        temp_process = temp_record.copy()
        # === 获取最大值和最小值
        min_value = temp_process[indicator].min()
        max_value = temp_process[indicator].max()
        # === 获取上下界标签
        min_group_label = temp_process.loc[temp_process[indicator] == min_value,indicator+"_K Means Label"].values[0]
        max_group_label = temp_process.loc[temp_process[indicator] == max_value,indicator+"_K Means Label"].values[0]
        # === 计算上下边界聚类
        lower_class = temp_process[temp_process[indicator+"_K Means Label"] == min_group_label]
        lower_class = np.array([float(i) for i in lower_class[indicator]])
        upper_class = temp_process[temp_process[indicator+"_K Means Label"] == max_group_label]
        upper_class = np.array([float(i) for i in upper_class[indicator]])
        
        # ***** 阈值算法 - 代码开始处*****
        # === 非加权算法
        overall_mean = np.mean(temp_process[indicator])
        overall_std = np.std(temp_process[indicator])

        # === 计算标准正态分布的分位数（给定概率计算)
        nd_quant_value, _= norm.interval(prob/100)
        nd_quant_value = abs(nd_quant_value)
        # === 计算上线边界
        if max_value / overall_mean > 40:
            upper_yellow_thres = (overall_mean) + nd_quant_value * (np.std(upper_class) + overall_std)
            upper_red_thres = upper_yellow_thres + (yellow_red_range + nd_quant_value) * (np.std(upper_class) + overall_std)
        else:
            upper_yellow_thres = (np.mean(upper_class) + overall_mean*2)/3 + nd_quant_value * (np.std(upper_class) + overall_std)
            upper_red_thres = upper_yellow_thres + (yellow_red_range + nd_quant_value) * (np.std(upper_class) + overall_std)
        # === 计算下线边界
        if max_value / overall_mean > 40:
            lower_yellow_thres = (np.mean(lower_class) + overall_mean*2)/3 - nd_quant_value * (np.std(lower_class) + overall_std)
            lower_red_thres = lower_yellow_thres - (yellow_red_range + nd_quant_value) * (np.std(lower_class) + overall_std)
        else:
            lower_yellow_thres = (np.mean(lower_class) + overall_mean*2)/3 - nd_quant_value * (np.std(lower_class) + overall_std)
            lower_red_thres = lower_yellow_thres - (yellow_red_range + nd_quant_value) * (np.std(lower_class) + overall_std)

        # === 对下线黄线做处理 - 不能大于0
        if lower_yellow_thres > 0:
            lower_yellow_thres = -0.1
        # === 将阈值扩展到百分比
        lower_yellow_thres = round(lower_yellow_thres,2)
        upper_yellow_thres = round(upper_yellow_thres,2)
        lower_red_thres = round(lower_red_thres,2)
        upper_red_thres = round(upper_red_thres,2)

        # === 所有数据都按照0个位数取整
        def Five_Zero(threshold):
            process_ts = int(round(threshold * 100))
            pos_neg_sign = 1 if threshold >= 0 else -1
            final_digit = int(str(process_ts)[-1])
            if final_digit == 1:
                return (process_ts - (1 * pos_neg_sign))/100
            elif final_digit == 2:
                return (process_ts - (2 * pos_neg_sign))/100
            elif final_digit == 3:
                return (process_ts - (3 * pos_neg_sign))/100
            elif final_digit == 4:
                return (process_ts - (4 * pos_neg_sign))/100
            elif final_digit == 5:
                return (process_ts + (5 * pos_neg_sign))/100
            elif final_digit == 6:
                return (process_ts + (4 * pos_neg_sign))/100
            elif final_digit == 7:
                return (process_ts + (3 * pos_neg_sign))/100
            elif final_digit == 8:
                return (process_ts + (2 * pos_neg_sign))/100
            elif final_digit == 9:
                return (process_ts + (1 * pos_neg_sign))/100
            else:
                return process_ts/100
        lower_yellow_thres = Five_Zero(lower_yellow_thres)
        upper_yellow_thres = Five_Zero(upper_yellow_thres)
        lower_red_thres = Five_Zero(lower_red_thres)
        upper_red_thres = Five_Zero(upper_red_thres)

        # === 对于小于0.05的值统一放在0.05
        if lower_yellow_thres >= -0.05:
            lower_yellow_thres = -0.05
            lower_red_thres = -0.1
        elif lower_yellow_thres >= -0.1:
            lower_yellow_thres = -0.1
            lower_red_thres = -0.15
        if upper_yellow_thres <= 0.05:
            upper_yellow_thres = 0.05
            upper_red_thres = 0.1
        elif upper_yellow_thres <= 0.1:
            upper_yellow_thres = 0.1
            upper_red_thres = 0.15
        
        # === 将相近结果区分开
        if abs(upper_red_thres - upper_yellow_thres) < 0.01:
            upper_red_thres = upper_red_thres + 0.05
        if abs(lower_red_thres - lower_yellow_thres) < 1:
            lower_red_thres = lower_red_thres - 0.05
        
        # === 对下线红线与做处理 - 不能小于100%和大于黄线
        if lower_red_thres <= -1:
            lower_red_thres = -1.01
        if lower_yellow_thres <= lower_red_thres:
            lower_yellow_thres = -0.9
            
        # === 缺乏历史数据时设置初始化指标 (所有历史数据均为0)
        if np.std(temp_process[indicator]) == 0:
            lower_yellow_thres = -0.05
            lower_red_thres = -0.10
            upper_yellow_thres = 0.05
            upper_red_thres = 0.10
        
        # ***** 阈值算法 - 代码结束处*****        
        # === 获取边界击中次数并年化
        risk_hit = len([i for i in temp_record[indicator] if i <= lower_red_thres]) + len([i for i in temp_record[indicator] if i >= upper_red_thres])
        sus_hit = len([i for i in temp_record[indicator] if i <= lower_yellow_thres]) + len([i for i in temp_record[indicator] if i >= upper_yellow_thres]) - risk_hit
        if risk_hit != 0:
            annualize_risk_hit = round((risk_hit/len([i for i in temp_record[indicator]])) * 12,2)
        else:
            annualize_risk_hit = 0
        if sus_hit != 0:
            annualize_sus_hit = round((sus_hit/len([i for i in temp_record[indicator]])) * 12,2)
        else:
            annualize_sus_hit = 0

        # ***** 数据可视化 *****
        # === 可视化数据计算结果
        temp_df = temp_process.copy()
        temp_df["time"] = [i for i in range(0,len(temp_df))]
        Lower_yellow = Threshold_result.loc[Threshold_result["Indicator"]==indicator,"Lower_yellow"].values[0]
        Lower_red = Threshold_result.loc[Threshold_result["Indicator"]==indicator,"Lower_red"].values[0]
        Upper_yellow = Threshold_result.loc[Threshold_result["Indicator"]==indicator,"Upper_yellow"].values[0]
        Upper_red = Threshold_result.loc[Threshold_result["Indicator"]==indicator,"Upper_red"].values[0]
        temp_df["Lower_yellow"] = Lower_yellow
        temp_df["Lower_red"] = Lower_red
        temp_df["Upper_yellow"] = Upper_yellow
        temp_df["Upper_red"] = Upper_red
        # === 建立三幅图及比例确定
        ax_main = inset_axes(ax_left,width="70%",height="80%",loc=2)
        ax_right = inset_axes(ax_left,width="25%",height="95%",loc=1)
        ax_bottom = inset_axes(ax_left,width="70%",height="15%",loc=3)
        # === 设置背景色
        ax_right.set_facecolor(self.theme[self.theme_select][0])
        ax_bottom.set_facecolor(self.theme[self.theme_select][0])
        ax_main.set_facecolor(self.theme[self.theme_select][0])
        
        # === 对主图绘图
        ax_main.plot(temp_df["time"],temp_df[indicator],color=line_color, marker=".",label=indicator,linewidth=2)
        ax_main.plot(temp_df["time"],temp_df["Lower_yellow"],color=self.theme[self.theme_select][3],label="黄色下线预警线",linewidth=2)
        ax_main.plot(temp_df["time"],temp_df["Upper_yellow"],color=self.theme[self.theme_select][3],label="黄色上线预警线",linewidth=2)
        ax_main.plot(temp_df["time"],temp_df["Lower_red"],color=self.theme[self.theme_select][4],label="红色下线预警线",linewidth=2)
        ax_main.plot(temp_df["time"],temp_df["Upper_red"],color=self.theme[self.theme_select][4],label="红色上线预警线",linewidth=2)
        ax_main.grid(axis="y",linestyle="--")
        # === 显示threshold 数据点 - 在最后一位显示
        ax_main.text(temp_df["time"].to_list()[0],temp_df["Lower_yellow"].to_list()[0],str(round(temp_df["Lower_yellow"].to_list()[0]*100))+"%",ha="right",va="top",fontsize=10, color=self.theme[self.theme_select][5],bbox=dict(boxstyle="round",fc=self.theme[self.theme_select][7],edgecolor=self.theme[self.theme_select][8],alpha=0.9))
        ax_main.text(temp_df["time"].to_list()[0],temp_df["Upper_yellow"].to_list()[0],str(round(temp_df["Upper_yellow"].to_list()[0]*100))+"%",ha="right",va="bottom",fontsize=10, color=self.theme[self.theme_select][5],bbox=dict(boxstyle="round",fc=self.theme[self.theme_select][7],edgecolor=self.theme[self.theme_select][8],alpha=0.9))
        ax_main.text(temp_df["time"].to_list()[-1],temp_df["Lower_red"].to_list()[-1],str(round(temp_df["Lower_red"].to_list()[-1]*100))+"%",ha="right",va="bottom",fontsize=10,color=self.theme[self.theme_select][6],bbox=dict(boxstyle="round",fc=self.theme[self.theme_select][7],edgecolor=self.theme[self.theme_select][8],alpha=0.9))
        ax_main.text(temp_df["time"].to_list()[-1],temp_df["Upper_red"].to_list()[-1],str(round(temp_df["Upper_red"].to_list()[-1]*100))+"%",ha="right",va="top",fontsize=10,color=self.theme[self.theme_select][6],bbox=dict(boxstyle="round",fc=self.theme[self.theme_select][7],edgecolor=self.theme[self.theme_select][8],alpha=0.9))

        # === 显示数据点
        date_label = [i for i in temp_df["time"]]
        value_label = [i for i in temp_df[indicator]]
        value_true_label= [str(round(i*100,1))+"%" for i in temp_df[indicator]]
        for a,b,c in zip(date_label,value_label,value_true_label):
            ax_main.text(a,b,c,ha="right",va="bottom",fontsize=9,color=line_color)
        ax_main.yaxis.set_major_formatter(ticker.PercentFormatter(xmax=1,decimals=0))
        ax_main.set_fontsize = (12)
        ax_main.grid(linestyle = '--')
        ax_main.set_xticks([])
        # === 对右侧图绘图
        ax_right.scatter(temp_df["time"],temp_df[indicator],c=kmeans.labels_,cmap="Set2",s=50) # 透明度
        ax_right.set_fontsize = (12)
        ax_right.set_xticks([])
        ax_right.set_yticks([])
        # === 对底部图绘图
        ax_bottom.hist(temp_df[indicator],  color='cadetblue',rwidth=0.5)
        ax_bottom.set_fontsize = (12)
        x = np.linspace(Lower_red-0.2,Upper_red+0.2,1000)
        fx = (1 / (overall_std * (2 * np.pi)**0.5) * np.exp(-(x - overall_mean)**2 / (2 * overall_std**2)))*20
        ax_bottom.plot(x,fx,color=self.theme[self.theme_select][2])
        ax_bottom.axvline(Lower_yellow, color=self.theme[self.theme_select][5],linewidth=2)
        ax_bottom.axvline(Lower_red, color=self.theme[self.theme_select][6],linewidth=2)
        ax_bottom.axvline(Upper_yellow, color=self.theme[self.theme_select][5],linewidth=2)
        ax_bottom.axvline(Upper_red, color=self.theme[self.theme_select][6],linewidth=2)
        ax_bottom.set_xticks([])

        ax_main.spines["top"].set_color(self.theme[self.theme_select][1])
        ax_main.spines["left"].set_color(self.theme[self.theme_select][1])
        ax_main.spines["right"].set_color(self.theme[self.theme_select][1])
        ax_main.spines["bottom"].set_color(self.theme[self.theme_select][1])
        
        ax_right.spines["top"].set_color(self.theme[self.theme_select][1])
        ax_right.spines["left"].set_color(self.theme[self.theme_select][1])
        ax_right.spines["right"].set_color(self.theme[self.theme_select][1])
        ax_right.spines["bottom"].set_color(self.theme[self.theme_select][1])
        ax_bottom.spines["top"].set_color(self.theme[self.theme_select][1])
        ax_bottom.spines["left"].set_color(self.theme[self.theme_select][1])
        ax_bottom.spines["right"].set_color(self.theme[self.theme_select][1])
        ax_bottom.spines["bottom"].set_color(self.theme[self.theme_select][1])

        ax_main.tick_params(axis='x', colors=self.theme[self.theme_select][1])
        ax_main.tick_params(axis='y', colors=self.theme[self.theme_select][1])
        ax_right.tick_params(axis='x', colors=self.theme[self.theme_select][1])
        ax_right.tick_params(axis='y', colors=self.theme[self.theme_select][1])
        ax_bottom.tick_params(axis='x', colors=self.theme[self.theme_select][1])
        ax_bottom.tick_params(axis='y', colors=self.theme[self.theme_select][1])
        return Threshold_result, ax_main, temp_df, temp_df
    # === Function: 所有指标回溯绘图
    def All_Metreics_Resume(self,good_data,service_data,figure_canvas,n_cate,yellow_red_range,prob,result_data):
        # 对result_data进行数据预处理
        result_data["Indicator"] = result_data["Indicator"].apply(lambda x: x[x.find("-")+1:len(x)])
        result_data["classifier"] = result_data["Indicator"].apply(lambda x: x[0:4])
        result_data_trade = result_data[result_data["classifier"]=="货物贸易"]
        result_data_trade["Indicator"] = result_data_trade["Indicator"].apply(lambda x: x[x.find("-")+1:len(x)])
        del result_data_trade["classifier"]
        result_data_service = result_data[result_data["classifier"]=="服务贸易"]
        result_data_service["Indicator"] = result_data_service["Indicator"].apply(lambda x: x[x.find("-")+1:len(x)])
        del result_data_service["classifier"]

        # 创建变量
        self.hit_record = {}
        self.name_dict = {}
        name_count = 1
        # 清空所有子图
        figure_canvas.clear()
        # 创建变量 - 绘图定位数据记录
        self.visual_record = pd.DataFrame(columns=["indicator","row","col"])
        # 定义中文输出
        # self.amend_Ops_text("中文输出定义完成")
        plt.rcParams['font.sans-serif']=['Microsoft YaHei']
        plt.rcParams['axes.unicode_minus'] = False
        plt.clf()
        self.amend_Ops_text("Completed for cleaning canvas")
        self.amend_Ops_text("Start to train K-Means model for param setup")
        # === 构建结果列表
        result_trade = pd.DataFrame(columns = ["Indicator","Upper_red","Upper_yellow","Lower_yellow","Lower_red","Sus_hit","Risk_hit"])
        # ********************* 绘图函数 ***************************
        # 计算货物贸易与服务贸易的总共数量
        goods_col = good_data.columns.tolist()
        service_col = service_data.columns.tolist()
        # 规定每行显示的图表数量
        col_show = 4
        total_len = len(goods_col) + len(service_col) # 获取总长度
        # 计算所需要的列
        row_show = math.ceil(total_len / col_show)+12
        # 初始化row coordinate 和 col coordinate 以及grid 构图
        row_coordinate, col_coordinate = 0,0
        grid_panel = plt.GridSpec(row_show, col_show, hspace=0.5, wspace=0.2)
        # 计算进度
        each_step = (80-40)/total_len
        current_pro = 40
        # === 绘制货物贸易图
        if len(goods_col) > 0:
            for i in goods_col:
                # 选取对应的行
                cut_df = result_data_trade.loc[result_data_trade["Indicator"]==i]
                ax = figure_canvas.add_subplot(grid_panel[row_coordinate:row_coordinate+1, col_coordinate:col_coordinate+1],facecolor="black")
                ax.set_title("货物贸易： "+i + "\n历史波动及阈值设置推荐_Code: "+str(name_count),fontdict={"color":self.theme[self.theme_select][1]})
                ax.axis('off')
                temp,globals()["ax_main"+str(name_count)],globals()["temp_df"+str(name_count)],data_df = self.K_Means_Algorithm_Resume(indicator = i,ax_left=ax,n_cate=n_cate,yellow_red_range=yellow_red_range,prob=prob,source_data=good_data,line_color=self.theme[self.theme_select][2],Threshold_result=cut_df)
                self.name_dict[str(name_count)+"-货物贸易-" + i] = str(name_count)
                temp["Indicator"] = temp["Indicator"].apply(lambda x: str(name_count)+"-货物贸易-" + x)
                result_trade = pd.concat([result_trade,temp])
                # 最近一年的击中情况放入字典中
                latest_temp = data_df.copy().tail(12)
                latest_temp["hit"] = 0
                latest_temp["hit"] = latest_temp[i].apply(lambda x: 1 if x > [i for i in latest_temp["Upper_yellow"]][0] else 0)
                latest_temp["hit"] = latest_temp[i].apply(lambda x: 1 if x < [i for i in latest_temp["Lower_yellow"]][0] else 0)
                hit_list = [i for i in latest_temp["hit"]]
                if len(hit_list) < 12:
                    hit_list = (12-len(hit_list)) * [0] + hit_list
                self.hit_record[str(name_count)+"-货物贸易-"+i] = hit_list
                # 将绘图记录放入visual record中
                self.visual_record.loc[len(self.visual_record)+1] = [str(name_count)+"-货物贸易-" + i,row_coordinate,col_coordinate]
                name_count += 1
                # 行列重新计算 - 智能布局
                if col_coordinate < (col_show-1):
                    col_coordinate += 1
                elif col_coordinate == (col_show-1):
                    row_coordinate += 1
                    col_coordinate = 0
                self.amend_Ops_text("指标计算完成：货物贸易 - "+i)
                current_pro += each_step
                self.progress_update(current_pro)
        row_coordinate = row_coordinate + 1
        col_coordinate = 0
        # === 绘制服务贸易图
        if len(service_col) > 0:
            for i in service_col:
                # 选取对应的行
                cut_df = result_data_service.loc[result_data_service["Indicator"]==i]
                ax = figure_canvas.add_subplot(grid_panel[row_coordinate:row_coordinate+1, col_coordinate:col_coordinate+1])
                ax.set_title("服务贸易： "+i + "\n历史波动及阈值设置推荐_Code: "+str(name_count),fontdict={"color":self.theme[self.theme_select][1]})
                ax.axis('off')
                temp,globals()["ax_main"+str(name_count)],globals()["temp_df"+str(name_count)],data_df= self.K_Means_Algorithm_Resume(indicator = i,ax_left=ax,n_cate=n_cate,yellow_red_range=yellow_red_range,prob=prob,source_data=service_data,line_color=self.theme[self.theme_select][9],Threshold_result=cut_df)
                self.name_dict[str(name_count)+"-服务贸易-" + i] = str(name_count)
                temp["Indicator"] = temp["Indicator"].apply(lambda x: str(name_count)+"-服务贸易-" + x)
                result_trade = pd.concat([result_trade,temp])
                # 最近一年的击中情况放入字典中
                latest_temp = data_df.copy().tail(12)
                latest_temp["hit"] = 0
                latest_temp["hit"] = latest_temp[i].apply(lambda x: 1 if x > [i for i in latest_temp["Upper_yellow"]][0] else 0)
                latest_temp["hit"] = latest_temp[i].apply(lambda x: 1 if x < [i for i in latest_temp["Lower_yellow"]][0] else 0)
                hit_list = [i for i in latest_temp["hit"]]
                if len(hit_list) < 12:
                    hit_list = (12-len(hit_list)) * [0] + hit_list
                self.hit_record[str(name_count)+"-货物贸易-"+i] = hit_list
                # 将绘图记录放入visual record中
                self.visual_record.loc[len(self.visual_record)+1] = [str(name_count)+"-服务贸易-" + i,row_coordinate,col_coordinate]
                name_count += 1
                # 行列重新计算 - 智能布局
                if col_coordinate < (col_show-1):
                    col_coordinate += 1
                elif col_coordinate == (col_show-1):
                    row_coordinate += 1
                    col_coordinate = 0
                self.amend_Ops_text("指标计算完成：货物贸易 - "+i)
                current_pro += each_step
                self.progress_update(current_pro)
        self.amend_Ops_text("Compute the Heatmaps")
        # === 加入相关性矩阵图
        ax_hm_1 = figure_canvas.add_subplot(grid_panel[row_coordinate+1:row_coordinate+5, 0:4])
        ax_hm_1.set_title("Correlation Matrix Heat Map - Goods",fontdict={"color":self.theme[self.theme_select][1]})
        ax_hm_1.set_facecolor(self.theme[self.theme_select][0])
        ax_hm_1.tick_params(axis='x', colors=self.theme[self.theme_select][1])
        ax_hm_1.tick_params(axis='y', colors=self.theme[self.theme_select][1])
        used_col = []
        for i in self.heap_goods.columns.to_list():
            if self.heap_goods[i].sum()!=0:
                used_col.append(i)
        input_matrix_copy = self.heap_goods[used_col]
        if len(input_matrix_copy.columns.to_list()) > 2:
            input_matrix_copy = input_matrix_copy.fillna(0)
            input_matrix_copy = input_matrix_copy.corr()
            sns.heatmap(input_matrix_copy, annot=True, vmax=1, square=True, cmap=self.theme[self.theme_select][10],ax=ax_hm_1)
        
        # === 加入相关性矩阵图
        ax_hm_2 = figure_canvas.add_subplot(grid_panel[row_coordinate+6:row_coordinate+10, 0:4])
        ax_hm_2.set_title("Correlation Matrix Heat Map - Services",fontdict={"color":self.theme[self.theme_select][1]})
        ax_hm_2.set_facecolor(self.theme[self.theme_select][0])
        ax_hm_2.tick_params(axis='x', colors=self.theme[self.theme_select][1])
        ax_hm_2.tick_params(axis='y', colors=self.theme[self.theme_select][1])
        used_col = []
        for i in self.heap_service.columns.to_list():
            if self.heap_service[i].sum()!=0:
                used_col.append(i)
        input_matrix_copy = self.heap_service[used_col]
        if len(input_matrix_copy.columns.to_list()) > 2:
            input_matrix_copy = input_matrix_copy.fillna(0)
            input_matrix_copy = input_matrix_copy.corr()
            sns.heatmap(input_matrix_copy, annot=True, vmax=1, square=True, cmap=self.theme[self.theme_select][10],ax=ax_hm_2)

        # === Canvas 绘图
        self.Latest_Hit_Count(self.hit_record)
        figure_canvas.suptitle("Whitelist Threshold Setup Recommendation Data Panel\n\n\n"+self.chn_name+"\n\n\n Scroll Down to View Visualized Result\n\n"+\
                                     '1 Year Hit Month Count: '+str(12 - self.hit_outcome), fontsize=26, color = self.theme[self.theme_select][1])
        self.output_fig = figure_canvas
        height_para = row_show * 450
        ini_canvas = tk.Canvas(self.window_2,width=2000,height=1150,scrollregion=(0,0,2100,height_para))
        ini_canvas.pack(fill=BOTH)
        dashboard = tk.Frame(ini_canvas)
        vbar=Scrollbar(ini_canvas,orient=VERTICAL)
        vbar.place(x=2000,height=1150,width=40)
        vbar.configure(command=ini_canvas.yview)
        ini_canvas.config(yscrollcommand=vbar.set)   
        ini_canvas.create_window((2000/2,height_para/2),window=dashboard,height=height_para,width=2000)
        self.canvas = FigureCanvasTkAgg(figure_canvas, master=dashboard) 
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        # 返回结果
        return result_trade
    # === Function: 依据前端进行绘图修改
    def Visualization_Amendment(self,indicator,threshold_list):
        # 1.获取Data Panel并清空数据
        ax_main = eval("ax_main"+self.name_dict[indicator])
        temp_df = eval("temp_df"+self.name_dict[indicator])
        ax_main.clear()
        # 2. 将新的指标阈值加入替换temp_df
        temp_df["Lower_yellow"] = threshold_list[0]
        temp_df["Lower_red"] = threshold_list[1]
        temp_df["Upper_yellow"] = threshold_list[2]
        temp_df["Upper_red"] = threshold_list[3]
        # 3. 重新绘制图表
        ori_indicator = indicator[indicator.find("-")+6:len(indicator)]
        # line color
        line_color = self.theme[self.theme_select][2] if "货物" in indicator else self.theme[self.theme_select][9]
        ax_main.plot(temp_df["time"],temp_df[ori_indicator],color=line_color, marker=".",label=indicator,linewidth=2)
        ax_main.plot(temp_df["time"],temp_df["Lower_yellow"],color=self.theme[self.theme_select][3],label="黄色下线预警线",linewidth=2)
        ax_main.plot(temp_df["time"],temp_df["Upper_yellow"],color=self.theme[self.theme_select][3],label="黄色上线预警线",linewidth=2)
        ax_main.plot(temp_df["time"],temp_df["Lower_red"],color=self.theme[self.theme_select][4],label="红色下线预警线",linewidth=2)
        ax_main.plot(temp_df["time"],temp_df["Upper_red"],color=self.theme[self.theme_select][4],label="红色上线预警线",linewidth=2)
        
        # === 显示threshold 数据点 - 在最后一位显示
        ax_main.text(temp_df["time"].to_list()[0],temp_df["Lower_yellow"].to_list()[0],str(round(temp_df["Lower_yellow"].to_list()[0]*100))+"%",ha="right",va="top",fontsize=10, color=self.theme[self.theme_select][5],bbox=dict(boxstyle="round",fc=self.theme[self.theme_select][7],edgecolor=self.theme[self.theme_select][8],alpha=0.9))
        ax_main.text(temp_df["time"].to_list()[0],temp_df["Upper_yellow"].to_list()[0],str(round(temp_df["Upper_yellow"].to_list()[0]*100))+"%",ha="right",va="bottom",fontsize=10, color=self.theme[self.theme_select][5],bbox=dict(boxstyle="round",fc=self.theme[self.theme_select][7],edgecolor=self.theme[self.theme_select][8],alpha=0.9))
        ax_main.text(temp_df["time"].to_list()[-1],temp_df["Lower_red"].to_list()[-1],str(round(temp_df["Lower_red"].to_list()[-1]*100))+"%",ha="right",va="bottom",fontsize=10,color=self.theme[self.theme_select][6],bbox=dict(boxstyle="round",fc=self.theme[self.theme_select][7],edgecolor=self.theme[self.theme_select][8],alpha=0.9))
        ax_main.text(temp_df["time"].to_list()[-1],temp_df["Upper_red"].to_list()[-1],str(round(temp_df["Upper_red"].to_list()[-1]*100))+"%",ha="right",va="top",fontsize=10,color=self.theme[self.theme_select][6],bbox=dict(boxstyle="round",fc=self.theme[self.theme_select][7],edgecolor=self.theme[self.theme_select][8],alpha=0.9))

        # === 显示数据点
        date_label = [i for i in temp_df["time"]]
        value_label = [i for i in temp_df[ori_indicator]]
        value_true_label= [str(round(i*100,1))+"%" for i in temp_df[ori_indicator]]
        for a,b,c in zip(date_label,value_label,value_true_label):
            ax_main.text(a,b,c,ha="right",va="bottom",fontsize=9,color=line_color)
        ax_main.yaxis.set_major_formatter(ticker.PercentFormatter(xmax=1,decimals=0))
        ax_main.set_fontsize = (12)
        ax_main.grid(linestyle = '--')
        ax_main.set_xticks([])
        self.canvas.draw()
        self.window_2.update()
        # === 修改获取边界击中次数并年化
        risk_hit = len([i for i in temp_df[ori_indicator] if i <= threshold_list[1]]) + len([i for i in temp_df[ori_indicator] if i >= threshold_list[3]])
        sus_hit = len([i for i in temp_df[ori_indicator] if i <= threshold_list[0]]) + len([i for i in temp_df[ori_indicator] if i >= threshold_list[2]]) - risk_hit
        if risk_hit != 0:
            annualize_risk_hit = round((risk_hit/len([i for i in temp_df[ori_indicator]])) * 12,2)
        else:
            annualize_risk_hit = 0
        if sus_hit != 0:
            annualize_sus_hit = round((sus_hit/len([i for i in temp_df[ori_indicator]])) * 12,2)
        else:
            annualize_sus_hit = 0
        
        # 最近一年的击中情况放入字典中
        latest_temp = temp_df.copy().tail(12)
        latest_temp["hit"] = 0
        latest_temp["hit"] = latest_temp[ori_indicator].apply(lambda x: 1 if x > [i for i in latest_temp["Upper_yellow"]][0] else 0)
        latest_temp["hit"] = latest_temp[ori_indicator].apply(lambda x: 1 if x < [i for i in latest_temp["Lower_yellow"]][0] else 0)
        hit_list = [i for i in latest_temp["hit"]]
        if len(hit_list) < 12:
            hit_list = (12-len(hit_list)) * [0] + hit_list
        self.hit_record[indicator] = hit_list
        self.Latest_Hit_Count(self.hit_record)
        self.fig_dashboard.suptitle("Whitelist Threshold Setup Recommendation Data Panel\n\n\n"+self.chn_name+"\n\n\n Scroll Down to View Visualized Result\n\n"+\
                                     '1 Year Hit Month Count: '+str(12 - self.hit_outcome), fontsize=26, color = self.theme[self.theme_select][1])
        self.window_2.update()
        return annualize_risk_hit, annualize_sus_hit
    # === Function: 计算最近一年的击中次数月份
    def Latest_Hit_Count(self,hit_dict):
        hit_data = pd.DataFrame(hit_dict)
        hit_data = hit_data.fillna(0)
        hit_data["outcome"] = hit_data[hit_data.columns.to_list()].sum(axis=1)
        self.hit_outcome = [i for i in hit_data["outcome"]].count(0)
        label_docxfile = tk.Label(self.window, text="     ", font=(self.font, 13))
        label_docxfile.place(x=1200, y=913)
        label_docxfile = tk.Label(self.window, text=str(12 - self.hit_outcome), font=(self.font, 13))
        label_docxfile.place(x=1200, y=913)
        self.window.update()

    # ******************* GUI Bridge Functions *******************
    # === Back-End Function: 选择source_data
    def B_Function_source_data_select(self):
        self.progress_update(0)
        self.amend_Ops_text("正在导入数据路径")
        self.progress_update(50)
        self.source_data = askdirectory()
        file_list = os.listdir(self.source_data)
        for file in file_list:
            wb = load_workbook(self.source_data+"/"+file)
            ws = wb.worksheets[0]
            self.chn_name = ws["B3"].value
            self.org_code = ws["B1"].value
            signal = ws["A4"].value
            if "货物" in signal:
                self.goods_path = self.source_data+"/"+file
            else:
                self.service_path = self.source_data+"/"+file
        self.progress_update(60)
        self.amend_Ops_text("数据路径对象获取完成，进行数据预处理和矩阵变换")
        # === 调用dataset load，获取数据集
        self.goods,self.goods_ratio,self.service,self.service_ratio = self.data_import(self.goods_path,self.service_path)
        self.progress_update(80)
        # === 进度更新
        self.progress_update(100)
        self.window.update()
        self.amend_Ops_text("数据加载完成")
    # === Back-End Function: 创建弹窗并显示参数设置
    def B_Function_Parameter_Panel(self):
        if self.param_set == 1:
            self.amend_Ops_text("进入参数设置面板")
            # === Label - K-Means Category No.
            self.para_label_1 = tk.Label(self.window, text='K-Means Clustering No. K均值聚类数量', font=(self.font, 12))
            self.para_label_1.place(x=50, y=400)
            # === Entry
            self.k_no = ttk.Entry(self.window, width=23, bootstyle="danger")
            self.k_no.insert(0,str(self.k_no_value))
            self.k_no.place(x=1150, y=400,height=31)
            # === Lable - Normal Distribution Confidence Interval
            self.para_label_2 = tk.Label(self.window, text='Normal Distribution Confidence Interval 正态分布置信区间', font=(self.font, 12))
            self.para_label_2.place(x=50, y=450)
            # === Entry
            self.ci = ttk.Entry(self.window, width=23, bootstyle="danger")
            self.ci.insert(0,str(self.ci_value))
            self.ci.place(x=1150, y=450,height=31)
            # === Lable - Yellow Red Threshold Range
            self.para_label_3 = tk.Label(self.window, text='Yellow Red Threshold Multiplier 红绿灯阈值乘数', font=(self.font, 12))
            self.para_label_3.place(x=50, y=500)
            # === Entry
            self.yr_range = ttk.Entry(self.window, width=23, bootstyle="danger")
            self.yr_range.insert(0,str(self.yr_range_value))
            self.yr_range.place(x=1150, y=500,height=31)
            # === Button - save 保存参数设置
            def save_parameters():
                self.k_no_value = int(self.k_no.get())
                self.ci_value = float(self.ci.get())
                self.yr_range_value = float(self.yr_range.get())
                self.amend_Ops_text("参数设置保存成功")
            self.btn_save = ttk.Button(self.window, text="Save Setup", command = save_parameters, bootstyle="danger")
            self. btn_save.place(x=1150, y=550, width=150)
            self.tree.place_forget()
            self.param_set = 0
        elif self.param_set == 0:
            self.amend_Ops_text("退出参数设置面板")
            self.btn_save.place_forget()
            self.para_label_1.place_forget()
            self.para_label_2.place_forget()
            self.para_label_3.place_forget()
            self.yr_range.place_forget()
            self.k_no.place_forget()
            self.ci.place_forget()
            self.tree.delete(*self.tree.get_children())
            self.tree.place(x=50,y=400,width=1300,height=500)
            self.param_set = 1
    # === Back-End Function: 运行计算
    def B_Function_Compuations(self):
        if datetime.now() <= pd.to_datetime("2024-01-18"):
            self.theme_select = self.theme_select.get()
            self.theme_select = "White" if self.theme_select == 0 else "Black"
            self.fig_dashboard.set_facecolor(self.theme[self.theme_select][0])
            self.amend_Ops_text("初始化表格内容完毕")
            self.progress_update(0)
            self.tree.delete(*self.tree.get_children())
            self.amend_Ops_text("数据容器创建完成")
            self.progress_update(20)
            self.result = pd.DataFrame(columns = ["Indicator","Upper_red","Upper_yellow","Lower_yellow","Lower_red","c","Risk_hit"])
            self.progress_update(30)
            self.window.update()
            self.amend_Ops_text("将数据传入算法函数中进行计算")
            self.progress_update(40)
            # === 获取computation mode 进行计算
            goods_data = self.goods_ratio
            service_data = self.service_ratio
            # === 调用函数完成指标计算
            threshold_result = self.All_Metrics_Computations(good_data = goods_data,service_data = service_data,
                                                            figure_canvas = self.fig_dashboard,
                                                            n_cate = self.k_no_value, yellow_red_range = self.yr_range_value, prob=self.ci_value,)
            self.progress_update(80)
            self.amend_Ops_text("阈值计算训练完成")
            self.result = pd.concat([self.result,threshold_result])
            self.amend_Ops_text("数值拼接完成")
            self.progress_update(90)
            self.amend_Ops_text("数值去重完成")
            self.result.drop_duplicates(inplace=True,keep="last",subset=["Indicator"])
            self.result = self.result.iloc[::-1]
            self.progress_update(95)
            # ***** 将数据同步到前端Panel中
            self.amend_Ops_text("正在将数据同步到前端UI界面")
            if len(self.result) > 0:
                for i in range(0,len(self.result)):
                    self.tree.insert("",0,text=str(i) ,values=(str(self.result.iloc[i]["Indicator"]),
                                                            str(round(self.result.iloc[i]["Upper_red"]*100,5))+"%",
                                                            str(round(self.result.iloc[i]["Upper_yellow"]*100,5))+"%",
                                                            str(round(self.result.iloc[i]["Lower_yellow"]*100,5))+"%",
                                                            str(round(self.result.iloc[i]["Lower_red"]*100,5))+"%",
                                                            str(self.result.iloc[i]["Sus_hit"]),
                                                            str(self.result.iloc[i]["Risk_hit"])))
            self.amend_Ops_text("阈值计算完成")
            self.progress_update(90)
            self.B_Function_Export()
            self.progress_update(100)
            self.window.update()
        else:
            None
    # === Back-End Function: 读入JSON文件，导入指标设置
    def B_Function_Resume(self):
        self.theme_select = self.theme_select.get()
        self.theme_select = "White" if self.theme_select == 0 else "Black"
        self.fig_dashboard.set_facecolor(self.theme[self.theme_select][0])
        self.result = pd.DataFrame(columns = ["Indicator","Upper_red","Upper_yellow","Lower_yellow","Lower_red","c","Risk_hit"])
        # === 获取配置文件，传递至self.result
        self.progress_update(0)
        self.amend_Ops_text("正在导入指标设置JSON文件")
        self.progress_update(20)
        json_data = askopenfilename()
        self.progress_update(30)
        self.amend_Ops_text("数据路径对象获取完成，正在回溯指标设置")
        result_data = pd.DataFrame(pd.read_json(json_data))
        # === 调用函数完成指标计算
        self.progress_update(40)
        threshold_result = self.All_Metreics_Resume(good_data = self.goods_ratio,service_data = self.service_ratio,
                                                        figure_canvas = self.fig_dashboard,
                                                        n_cate = self.k_no_value, yellow_red_range = self.yr_range_value, prob=self.ci_value,result_data=result_data)
        self.progress_update(80)
        self.amend_Ops_text("阈值计算训练完成")
        self.result = pd.concat([self.result,threshold_result])
        self.amend_Ops_text("数值拼接完成")
        self.progress_update(90)
        self.amend_Ops_text("数值去重完成")
        self.result.drop_duplicates(inplace=True,keep="last",subset=["Indicator"])
        self.result = self.result.iloc[::-1]
        self.progress_update(95)
        # ***** 将数据同步到前端Panel中
        self.amend_Ops_text("正在将数据同步到前端UI界面")
        if len(self.result) > 0:
            for i in range(0,len(self.result)):
                self.tree.insert("",0,text=str(i) ,values=(str(self.result.iloc[i]["Indicator"]),
                                                        str(round(self.result.iloc[i]["Upper_red"]*100,5))+"%",
                                                        str(round(self.result.iloc[i]["Upper_yellow"]*100,5))+"%",
                                                        str(round(self.result.iloc[i]["Lower_yellow"]*100,5))+"%",
                                                        str(round(self.result.iloc[i]["Lower_red"]*100,5))+"%",
                                                        str(self.result.iloc[i]["Sus_hit"]),
                                                        str(self.result.iloc[i]["Risk_hit"])))
        self.amend_Ops_text("阈值计算完成")
        os.remove(json_data) # 先删除旧版本的json数据
        self.B_Function_Export()
        self.progress_update(100)
        self.window.update()
    # === Back-End Funtions: 保存本地结果
    def B_Function_Export(self):
            # === 创建一个文件夹
            output_path = self.source_data[0:self.source_data.rfind("/")]+"/"+self.org_code+"_Metrics_Setup_Outcome"
            if os.path.exists(output_path)==False:
                os.makedirs(output_path)
            else:
                None
            # === 结果表格预处理
            # === 1.0 生成Client Confirmation Table
            client_confirm = self.result.copy()
            # 对result_data进行数据预处理
            # 计算调整系数            
            client_confirm["正常区间A"] = client_confirm["Lower_yellow"].apply(lambda x: str(round(x*100,2))+"% ~ ")
            client_confirm["正常区间B"] = client_confirm["Upper_yellow"].apply(lambda x: str(round(x*100,2))+"%")
            client_confirm["正常区间"] = client_confirm["正常区间A"] + client_confirm["正常区间B"]
            client_confirm["关注区间A"] = client_confirm["Lower_red"].apply(lambda x: str(round(x*100,2))+"% ~ ")
            client_confirm["关注区间B"] = client_confirm["Lower_yellow"].apply(lambda x: str(round(x*100,2))+"%; ")
            client_confirm["关注区间C"] = client_confirm["Upper_yellow"].apply(lambda x: str(round(x*100,2))+"% ~ ")
            client_confirm["关注区间D"] = client_confirm["Upper_red"].apply(lambda x: str(round(x*100,2))+"%")
            client_confirm["关注区间"] = client_confirm["关注区间A"] + client_confirm["关注区间B"] + client_confirm["关注区间C"] + client_confirm["关注区间D"]
            client_confirm["可疑区间A"] = client_confirm["Lower_red"].apply(lambda x: "<"+str(round(x*100,2))+"%; ")
            client_confirm["可疑区间B"] = client_confirm["Upper_red"].apply(lambda x: ">"+str(round(x*100,2))+"%")
            client_confirm["可疑区间"] = client_confirm["可疑区间A"] + client_confirm["可疑区间B"]
            client_confirm = client_confirm[["Indicator","正常区间","关注区间","可疑区间"]]
            client_confirm = client_confirm.rename(columns={"Indicator":self.chn_name})
            
            goods_header = pd.DataFrame(columns=["货物贸易监测指标（月度/年度二选一，如选年度则无需设置同比指标）"])
            service_header = pd.DataFrame(columns=["服务贸易监测指标（月度/年度二选一，如选年度则无需设置同比指标）"])
            Basic_info = pd.DataFrame(columns=["Items","Values"])
            Basic_info["Items"] = ["指标设定时间","指标设定人","指标设定审批人","指标设定客户确认时间","指标设定客户确认记录"]
            Basic_info["Values"] = [None,None,None,None,None]

            # Data procesing
            client_confirm[self.chn_name] = client_confirm[self.chn_name].apply(lambda x: x[x.find("-")+1:len(x)])
            client_confirm["classifier"] = client_confirm[self.chn_name].apply(lambda x: x[0:4])
            client_confirm_trade = client_confirm[client_confirm["classifier"]=="货物贸易"]
            client_confirm_trade[self.chn_name] = client_confirm_trade[self.chn_name].apply(lambda x: x[x.find("-")+1:len(x)])
            del client_confirm_trade["classifier"]
            client_confirm_service = client_confirm[client_confirm["classifier"]=="服务贸易"]
            print(client_confirm_service)
            client_confirm_service[self.chn_name] = client_confirm_service[self.chn_name].apply(lambda x: x[x.find("-")+1:len(x)])
            del client_confirm_service["classifier"]
            # === 保存图片
            self.output_fig.savefig(output_path+"/"+str(self.org_code)+"_Param_Setup.jpg")

            Overview_output_path = output_path+"/"+str(self.org_code)+"_Client_Confirmation_Sheet.xlsx"
            with pd.ExcelWriter(Overview_output_path, mode="w",engine="xlsxwriter") as writer:
                Basic_info.to_excel(writer,sheet_name="Sheet1",startrow=0, index=False,header=False)
                goods_header.to_excel(writer,sheet_name="Sheet1",startrow=8, index=False)
                client_confirm_trade.to_excel(writer,sheet_name="Sheet1",startrow=9, index=False)
                service_header.to_excel(writer,sheet_name="Sheet1",startrow=9+len(client_confirm_trade)+3, index=False)
                client_confirm_service.to_excel(writer,sheet_name="Sheet1",startrow=9+len(client_confirm_trade)+4, index=False)
                worksheet = writer.sheets["Sheet1"]
                # set column width
                worksheet.set_column(0, 0, 80)
                worksheet.set_column(1, 3, 30)

            # === 将所有threshold前面的数字去除
            output_data = self.result.copy()
            output_data["Indicator"] = output_data["Indicator"].apply(lambda x:x[x.find("-")+1:len(x)])
            result_indicators = [i for i in output_data["Indicator"]]
            goods_indicator = []
            service_indicator = []
            for item in result_indicators:
                if item.startswith("货物贸易"):
                    goods_indicator.append(item)
                elif item.startswith("服务贸易"):
                    service_indicator.append(item)

            # === 2.0 生成 Paramter Setup Form
            param_setup_table = result_indicators = output_data.copy()
            # 创建goods和service表
            goods_sys = pd.DataFrame(columns = ["监测指标","监测阈值","调整系数"])
            service_sys = pd.DataFrame(columns = ["监测指标","监测阈值","调整系数"])
            for i in range(0,len(param_setup_table)):
                if param_setup_table.iloc[i]["Indicator"].startswith("货物贸易"):
                    biggest_value = max([param_setup_table.iloc[i]["Upper_red"],abs(param_setup_table.iloc[i]["Lower_red"])])
                    scaling_factor = 1 * 10**(-(str(biggest_value).find(".")-1))
                    goods_sys.loc[len(goods_sys)+1] = [
                                                            param_setup_table.iloc[i]["Indicator"][5:len(param_setup_table.iloc[i]["Indicator"])],
                                                            str(round(param_setup_table.iloc[i]["Upper_red"]*scaling_factor,2))+","+
                                                            str(round(param_setup_table.iloc[i]["Upper_yellow"]*scaling_factor,2))+","+
                                                            str(round(param_setup_table.iloc[i]["Lower_yellow"]*scaling_factor,2))+","+
                                                            str(round(param_setup_table.iloc[i]["Lower_red"]*scaling_factor,2)),
                                                            scaling_factor if scaling_factor != 1 else None
                                                        ]
                else:
                    biggest_value = max([param_setup_table.iloc[i]["Upper_red"],abs(param_setup_table.iloc[i]["Lower_red"])])
                    scaling_factor = 1 * 10**(-(str(biggest_value).find(".")-1))
                    service_sys.loc[len(service_sys)+1] = [
                                                            param_setup_table.iloc[i]["Indicator"][5:len(param_setup_table.iloc[i]["Indicator"])],
                                                            str(round(param_setup_table.iloc[i]["Upper_red"]*scaling_factor,2))+","+
                                                            str(round(param_setup_table.iloc[i]["Upper_yellow"]*scaling_factor,2))+","+
                                                            str(round(param_setup_table.iloc[i]["Lower_yellow"]*scaling_factor,2))+","+
                                                            str(round(param_setup_table.iloc[i]["Lower_red"]*scaling_factor,2)),
                                                            scaling_factor if scaling_factor != 1 else None
                                                        ]

            Overview_output_path = output_path+"/"+str(self.org_code)+"_Param_Setup.xlsx"
            with pd.ExcelWriter(Overview_output_path, mode="w",engine="xlsxwriter") as writer:
                # === 货物贸易指标
                if self.good_frequency == "Month":
                    pd.DataFrame(columns=["贸易外汇收支便利化试点交易指标监测系数设置（月报）"]).to_excel(writer,sheet_name="货物贸易月报",startrow=0, index=False)
                    pd.DataFrame(columns=["企业名称",self.chn_name]).to_excel(writer,sheet_name="货物贸易月报",startrow=1, index=False)
                    pd.DataFrame(columns=["组织机构代码",self.org_code]).to_excel(writer,sheet_name="货物贸易月报",startrow=2, index=False)
                    pd.DataFrame(columns=["指标检测类型","货物贸易 Goods Trade 月度"]).to_excel(writer,sheet_name="货物贸易月报",startrow=3, index=False)
                    goods_sys.to_excel(writer,sheet_name="货物贸易月报",startrow=5, index=False)
                    # set column width
                    worksheet = writer.sheets["货物贸易月报"]
                    worksheet.set_column(0, 3, 60)
                else:
                    pd.DataFrame(columns=["贸易外汇收支便利化试点交易指标监测系数设置（年报）"]).to_excel(writer,sheet_name="货物贸易年报",startrow=0, index=False)
                    pd.DataFrame(columns=["企业名称",self.chn_name]).to_excel(writer,sheet_name="货物贸易年报",startrow=1, index=False)
                    pd.DataFrame(columns=["组织机构代码",self.org_code]).to_excel(writer,sheet_name="货物贸易年报",startrow=2, index=False)
                    pd.DataFrame(columns=["指标检测类型","货物贸易 Goods Trade 年度"]).to_excel(writer,sheet_name="货物贸易年报",startrow=3, index=False)
                    goods_sys.to_excel(writer,sheet_name="货物贸易年报",startrow=5, index=False)
                    # set column width
                    worksheet = writer.sheets["货物贸易年报"]
                    worksheet.set_column(0, 3, 60)
                # === 服务贸易指标
                if self.service_frequency == "Month":
                    pd.DataFrame(columns=["贸易外汇收支便利化试点交易指标监测系数设置（月报）"]).to_excel(writer,sheet_name="服务贸易月报",startrow=0, index=False)
                    pd.DataFrame(columns=["企业名称",self.org_code]).to_excel(writer,sheet_name="服务贸易月报",startrow=1, index=False)
                    pd.DataFrame(columns=["组织机构代码",self.org_code]).to_excel(writer,sheet_name="服务贸易月报",startrow=2, index=False)
                    pd.DataFrame(columns=["指标检测类型","服务贸易 Service Trade 月度"]).to_excel(writer,sheet_name="服务贸易月报",startrow=3, index=False)
                    service_sys.to_excel(writer,sheet_name="服务贸易月报",startrow=5, index=False)
                    # set column width
                    worksheet = writer.sheets["服务贸易月报"]
                    worksheet.set_column(0, 3, 60)
                else:
                    pd.DataFrame(columns=["贸易外汇收支便利化试点交易指标监测系数设置（年报）"]).to_excel(writer,sheet_name="服务贸易年报",startrow=0, index=False)
                    pd.DataFrame(columns=["企业名称",self.org_code]).to_excel(writer,sheet_name="服务贸易年报",startrow=1, index=False)
                    pd.DataFrame(columns=["组织机构代码",self.org_code]).to_excel(writer,sheet_name="服务贸易年报",startrow=2, index=False)
                    pd.DataFrame(columns=["指标检测类型","服务贸易 Service Trade 年度"]).to_excel(writer,sheet_name="服务贸易年报",startrow=3, index=False)
                    service_sys.to_excel(writer,sheet_name="服务贸易年报",startrow=5, index=False)
                    # set column width
                    worksheet = writer.sheets["服务贸易年报"]
                    worksheet.set_column(0, 3, 60)
            # === 保存图片
            self.output_fig.savefig(output_path+"/"+str(self.org_code)+"_Param_Setup.jpg")
            # === 导出指标设置结果到JSON 文件中
            param_data = self.result.copy()
            param_data.reset_index(drop=True, inplace=True)
            param_data.to_json(output_path+"/"+str(self.org_code)+"_Metrics_Setup.json")
            self.amend_Ops_text("指标设置JSON文件保存成功")
        # === Back-End Function: 将表格回传到Core Algorithm进行绘图修改
    # === Back-End Functions: 打开drill report
    def B_Function_Open_Drill(self):  
        order_A = '"'+self.goods_path.replace("/","\\")+'"'
        order_A = 'start excel.exe '+order_A
        os.system(order_A)
        order_B = '"'+self.service_path.replace("/","\\")+'"'
        order_B = 'start excel.exe '+order_B
        os.system(order_B)
    # ******************* Graphical User Interface *******************
    # === UI Function: 创建主界面和绘画界面
    def window_setup(self):
        self.window = ttk.Window(themename="darkly")
        self.font = 'Yu Gothic UI'
        self.window.geometry("1400x1020")  # 窗口大小
        # ==== 窗口标题
        self.window.title("WhiteList Monitoring Parameter Setup")
        label_title = tk.Label(self.window, text='WhiteList Monitoring Parameter Setup',font=("Yu Gothic UI", 20), height=5)
        label_title.pack()
        # === 创建绘图  dashboard
        self.window_2 = tk.Tk()
        self.window_2.geometry("2040x1150")
        self.window_2.wm_title("Visualization Dashboard")
    # === UI Function: 更新窗口界面
    def window_show(self):
        self.window_2.mainloop()
        self.window.mainloop()
        self.data_panel.mainloop()
    # === UI Function: 更新参数面板
    def window_parameter_panel(self):
        # === 创建绘图面板  self.fig_dashboard
        plt.rcParams['font.sans-serif']=['Microsoft YaHei']
        plt.rcParams['axes.unicode_minus'] = False
        self.fig_dashboard = plt.figure(figsize=(20, 10), dpi=85)
        # === Label- 参数设置
        label_docxfile = tk.Label(self.window, text='Whitelist Parameters', font=(self.font, 14))
        label_docxfile.place(x=50, y=133)
        # === Button - 上传BOP数据
        style_button = ttk.Style()
        style_button.configure("TButton", font=(self.font, 11), justify='center')
        path_select = ttk.Button(self.window, text="Import BOP Data", command = self.B_Function_source_data_select, bootstyle="warning-outline")
        path_select.place(x=50, y=181, width=200)
        self.window.update()
        # === Button
        style_button = ttk.Style()
        style_button.configure("TButton", font=(self.font, 11), justify='center')
        path_select = ttk.Button(self.window, text="Open BOP Data", command = self.B_Function_Open_Drill, bootstyle="info-outline")
        path_select.place(x=325, y=181, width=200)
        self.window.update()
        # === Button
        style_button = ttk.Style()
        style_button.configure("TButton", font=(self.font, 11), justify='center')
        path_select = ttk.Button(self.window, text="K-Means Paramter Tune", command = self.B_Function_Parameter_Panel, bootstyle="danger-outline")
        path_select.place(x=600, y=181, width=200)
        self.window.update()
        # === Button - Compute
        style_button = ttk.Style()
        style_button.configure("TButton", font=(self.font, 11), justify='center')
        path_select = ttk.Button(self.window, text="Computaion", command = self.B_Function_Compuations, bootstyle="success-outline")
        path_select.place(x=875, y=181, width=200)
        self.window.update()
        label_docxfile = tk.Label(self.window, text='Progress Update', font=(self.font, 13))
        label_docxfile.place(x=50, y=232)
        # === Button - Resume / Json
        style_button = ttk.Style()
        style_button.configure("TButton", font=(self.font, 11), justify='center')
        path_select = ttk.Button(self.window, text="Resume (Import JSON)", command = self.B_Function_Resume, bootstyle="warning-outline")
        path_select.place(x=1150, y=181, width=200)
        self.window.update()
        label_docxfile = tk.Label(self.window, text='Progress Update', font=(self.font, 13))
        label_docxfile.place(x=50, y=232)
        # === Progress - Bar
        self.bar = ttk.Progressbar(self.window,length=931,value=0,bootstyle="warning-striped")
        self.bar.place(x=205,y=238,height=20)
        self.bar["value"] = 0
        self.window.update()
        # Status - Show
        self.Form_text = tk.Text(self.window, width=134, height=5, state='disabled')
        self.Form_text.place(x=50, y=270)
        self.Form_text.config(state = 'normal')
        font_text = tf.Font(family="Yu Gothic UI", size=11)
        self.Form_text.insert("insert","[System Response] -> Hello! Welcome to use Whitelist Model."+"\n")
        self.Form_text.config(font=font_text,foreground="goldenrod")
        self.Form_text.see("end")
        self.Form_text.config(state = 'disabled')
        self.window.update()
        # Status - Meter
        self.meter_tool = ttk.Meter(bootstyle="success",subtextstyle="success",metersize=160,stripethickness=5,
                                    amountused=0,meterthickness=14,subtext="% Computation")
        self.meter_tool.place(x=1180, y=225)
        # 最近一年没有hit的次数
        label_docxfile = tk.Label(self.window, text='1Y Hit Month Count:', font=(self.font, 13))
        label_docxfile.place(x=1000, y=913)
        # 色系
        # 最近一年没有hit的次数
        label_docxfile = tk.Label(self.window, text='Visual Light Theme', font=(self.font, 13))
        label_docxfile.place(x=1000, y=960)
        label_docxfile = tk.Label(self.window, text='Visual Dark Theme', font=(self.font, 13))
        label_docxfile.place(x=1200, y=960)
        self.theme_select = IntVar()
        theme = ttk.Checkbutton(bootstyle="info-square-toggle",offvalue=0,onvalue=1,variable=self.theme_select)
        theme.place(x=1163, y=968)

        
        # === 分割线
        Separator(self.window, orient=HORIZONTAL).place(x=1, y=15, width=2000)
        Separator(self.window, orient=HORIZONTAL).place(x=240, y=150, width=1125)
        Separator(self.window, orient=HORIZONTAL).place(x=40, y=150, width=10)
        Separator(self.window, orient=VERTICAL).place(x=40, y=150, height=851)
        Separator(self.window, orient=VERTICAL).place(x=1365, y=150, height=851)
        Separator(self.window, orient=HORIZONTAL).place(x=40, y=390, width=1325)
        Separator(self.window, orient=HORIZONTAL).place(x=40, y=1000, width=1325)
    # === UI Function: 运行结果表格
    def window_outcome_table(self):
        # 设置表格控件
        self.tree = ttk.Treeview(self.window,show='headings')
        style_head = ttk.Style()
        style_head.configure("Treeview.Heading", font = ("Yu Gothic UI",11), rowheight=20)
        style_head.configure("Treeview", font = ("Yu Gothic UI",11), rowheight=20,foreground="lightseagreen")
        self.tree["columns"]=("Indicator","Risky-Pos","Sus-Pos","Sus-Neg","Risky-Neg","Sus-Hit","Risk-Hit")
        # 设置列
        self.tree.column("Indicator",width=250)
        self.tree.column("Risky-Pos",width=5)
        self.tree.column("Sus-Pos",width=5)
        self.tree.column("Sus-Neg",width=5)
        self.tree.column("Risky-Neg",width=5)
        self.tree.column("Sus-Hit",width=5)
        self.tree.column("Risk-Hit",width=3)
        # 设置heading
        self.tree.heading("Indicator",text="Indicator")
        self.tree.heading("Risky-Pos",text="Risky-Pos")
        self.tree.heading("Sus-Pos",text="Suss-Pos")
        self.tree.heading("Sus-Neg",text="Sus-Neg")
        self.tree.heading("Risky-Neg",text="Risky-Neg")
        self.tree.heading("Sus-Hit",text="Sus-Hit")
        self.tree.heading("Risk-Hit",text="Risk-Hit")
        self.tree.place(x=50,y=400,width=1300,height=500)
        # === 事件驱动 - 双击可以修改表格内数据
        self.tree.bind('<Double-1>', self.edit_table) # 双击左键进入编辑
        # === 显示编辑窗口 - 更新按钮
        self.amend_value_pos_risk = ttk.Entry(self.window, width=10, bootstyle="primary")
        self.amend_value_pos_sus = ttk.Entry(self.window, width=10, bootstyle="primary")
        self.amend_value_neg_sus = ttk.Entry(self.window, width=10, bootstyle="primary")
        self.amend_value_neg_risk = ttk.Entry(self.window, width=10, bootstyle="primary")
        self.amend_value_pos_risk.place(x=50, y=910,height=33)
        self.amend_value_pos_sus.place(x=150, y=910,height=33)
        self.amend_value_neg_sus.place(x=250, y=910,height=33)
        self.amend_value_neg_risk.place(x=350, y=910,height=33)
        
        
        style_button = ttk.Style()
        style_button.configure("TButton", font=(self.font, 11), justify='center',background="lightseagreen")
        confirm_btn = ttk.Button(self.window, text="Update Result", command = None, )
        confirm_btn.place(x=50, y=955, width=375, height=33)
        self.window.update()
    # === UI Function: 表格编辑
    def edit_table(self,event):
        for item in self.tree.selection():
            item_text = self.tree.item(item,"values")
        # 获取行和列
        column = self.tree.identify_column(event.x)# 列
        # 获取鼠标双击的所在列
        # self.column_loc = int(column.replace("#",""))-1
        risk_pos = float(item_text[1].replace("%",""))
        sus_pos = float(item_text[2].replace("%",""))
        sus_neg = float(item_text[3].replace("%",""))
        risk_neg = float(item_text[4].replace("%",""))
        
        
        # 获取鼠标双击的indicator名称
        self.indicator_label = item_text[0]
        # 在treeview中加入数据
        self.amend_value_pos_risk.delete(0,END)
        self.amend_value_pos_risk.insert(0,risk_pos)
        self.amend_value_pos_sus.delete(0,END)
        self.amend_value_pos_sus.insert(0,sus_pos)
        self.amend_value_neg_sus.delete(0,END)
        self.amend_value_neg_sus.insert(0,sus_neg)
        self.amend_value_neg_risk.delete(0,END)
        self.amend_value_neg_risk.insert(0,risk_neg)
        
        # Internal Function：保存结果，并更新到treeview，result panel和图表修改
        def saveedit():
            # 将结果更新到UI界面中
            self.tree.set(item, column=1, value=str(round(float(self.amend_value_pos_risk.get()),5))+"%")
            self.tree.set(item, column=2, value=str(round(float(self.amend_value_pos_sus.get()),5))+"%")
            self.tree.set(item, column=3, value=str(round(float(self.amend_value_neg_sus.get()),5))+"%")
            self.tree.set(item, column=4, value=str(round(float(self.amend_value_neg_risk.get()),5))+"%")
            # 更新self.result的内容
            self.result.loc[self.result["Indicator"]==self.indicator_label,"Upper_red"] = float(self.amend_value_pos_risk.get())/100
            self.result.loc[self.result["Indicator"]==self.indicator_label,"Upper_yellow"] = float(self.amend_value_pos_sus.get())/100
            self.result.loc[self.result["Indicator"]==self.indicator_label,"Lower_yellow"] = float(self.amend_value_neg_sus.get())/100
            self.result.loc[self.result["Indicator"]==self.indicator_label,"Lower_red"] = float(self.amend_value_neg_risk.get())/100
            # 获取最新的threshold list
            thres_list = [float(self.result.loc[self.result["Indicator"]==self.indicator_label,"Lower_yellow"].values[0]),
                          float(self.result.loc[self.result["Indicator"]==self.indicator_label,"Lower_red"].values[0]),
                          float(self.result.loc[self.result["Indicator"]==self.indicator_label,"Upper_yellow"].values[0]),
                          float(self.result.loc[self.result["Indicator"]==self.indicator_label,"Upper_red"].values[0])]
            annualize_risk_hit, annualize_sus_hit = self.Visualization_Amendment(self.indicator_label,thres_list)
            # 更新threshold list
            self.result.loc[self.result["Indicator"]==self.indicator_label,"Risk_hit"] = annualize_risk_hit
            self.result.loc[self.result["Indicator"]==self.indicator_label,"Sus_hit"] = annualize_sus_hit
            # 反馈到前端UI面板并保存到本地
            self.tree.set(item, column=6, value=str(annualize_risk_hit))
            self.tree.set(item, column=5, value=str(annualize_sus_hit))
            self.amend_Ops_text("阈值更新成功,正在保存更新结果")
            self.B_Function_Export()
            self.amend_Ops_text("更新结果保存成功")
            
        #  Button - 调用函数
        style_button = ttk.Style()
        style_button.configure("TButton", font=(self.font, 11), justify='center',background="lightseagreen")
        confirm_btn = ttk.Button(self.window, text="Update Result", command = saveedit)
        confirm_btn.place(x=50, y=955, width=375, height=33)
        self.window.update()
    # === UI Function: 显示运行状态
    def amend_Ops_text(self, content):
        self.Form_text.config(state = 'normal')
        font_text = tf.Font(family="Yu Gothic UI", size=11)
        self.Form_text.insert("insert","[System Response] -> "+content+"\n")
        self.Form_text.config(font=font_text,foreground="goldenrod")
        self.Form_text.see("end")
        self.Form_text.config(state = 'disabled')
        self.window.update()
        return None
    # === UI Function: 显示状态栏
    def progress_update(self,status):
        status = round(status,1)
        self.meter_tool.configure(amountused = status)
        self.bar["value"] = status
        self.window.update()

    # ******************* Overall Execution *******************
    def Overall_Execution(self):
        self.window_setup()
        self.window_parameter_panel()
        self.window_outcome_table()
        self.window_show()


# ***** Main System Access *****
if __name__ == "__main__":
    Raphael = Threshold_Algorithm()
    Raphael.Overall_Execution()